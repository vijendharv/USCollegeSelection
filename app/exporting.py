"""Render PDF and Excel files exclusively from the canonical college report."""

from __future__ import annotations

from decimal import Decimal
from html import escape
from io import BytesIO
from typing import Any
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    TableStyle,
)
from reportlab.platypus import (
    Table as PDFTable,
)

from app.models.export import ExportedFile, ExportFormat, ExportResult
from app.models.report import CollegeReport, SchoolReport
from app.storage.contracts import SessionFileStore

_NAVY = "17365D"
_BLUE = "D9EAF7"
_LIGHT = "F4F7FA"
_WHITE = "FFFFFF"
_GREEN = "E2F0D9"
_YELLOW = "FFF2CC"
_RED = "FCE4D6"
_GRAY = "E7E6E6"
_THIN_GRAY = Side(style="thin", color="D9E1F2")
_GPA_UNAVAILABLE = "Not available from current official dataset"


def render_excel(report: CollegeReport) -> bytes:
    """Create the report workbook in memory."""
    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is not None:
        workbook.remove(active_sheet)
    if report.student_profile.preferences.existing_schools:
        _student_supplied_colleges_sheet(workbook, report)
    _college_list_sheet(workbook, report)
    _major_rankings_sheet(workbook, report)
    _adaptive_thresholds_sheet(workbook, report)
    _gap_analysis_sheet(workbook, report)
    _student_profile_sheet(workbook, report)
    _application_tracker_sheet(workbook, report)
    _sources_sheet(workbook, report)
    if report.addendum_rankings:
        _addendum_sheet(workbook, report)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def render_pdf(report: CollegeReport) -> bytes:
    """Create a printable PDF from the same canonical values."""
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=letter,
        rightMargin=0.55 * inch,
        leftMargin=0.55 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.55 * inch,
        title="US College Selection Report",
        author="US College Selection",
        pageCompression=1,
    )
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="ReportTitle",
            parent=styles["Title"],
            textColor=colors.HexColor(f"#{_NAVY}"),
            fontSize=19,
            leading=23,
            alignment=TA_CENTER,
            spaceAfter=10,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SchoolTitle",
            parent=styles["Heading1"],
            textColor=colors.HexColor(f"#{_NAVY}"),
            fontSize=14,
            leading=17,
            spaceAfter=6,
        )
    )
    styles.add(
        ParagraphStyle(
            name="Small",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
        )
    )
    styles.add(
        ParagraphStyle(
            name="TableHeader",
            parent=styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=7,
            leading=8,
            textColor=colors.white,
        )
    )
    story: list[Any] = [
        Paragraph("US College Selection Report", styles["ReportTitle"]),
        Paragraph(
            f"Generated {_pdf_text(report.generated_at.isoformat())} | "
            f"Methodology {_pdf_text(report.methodology_version)} | "
            f"Dataset {_pdf_text(report.dataset.version_id)}",
            styles["Small"],
        ),
        Spacer(1, 10),
        _pdf_summary(report, styles),
        Spacer(1, 12),
        Paragraph(_pdf_text(report.disclaimer), styles["Small"]),
    ]
    if report.data_quality_warnings:
        story.append(Spacer(1, 6))
        story.extend(
            Paragraph(f"Data note: {_pdf_text(value)}", styles["Small"])
            for value in report.data_quality_warnings
        )
    if report.program_data_vintages:
        story.append(
            Paragraph(
                "Program data: " + _pdf_text("; ".join(report.program_data_vintages)),
                styles["Small"],
            )
        )
    story.extend(_pdf_thresholds(report, styles))
    story.extend(_pdf_student_supplied_colleges(report, styles))
    story.append(PageBreak())
    story.extend(_pdf_major_ranking_sections(report, styles))
    if report.major_rankings:
        story.append(PageBreak())
    for index, school in enumerate(report.schools):
        category = school.classification.category.value.replace("_", " ").title()
        story.extend(
            [
                Paragraph(_pdf_text(school.institution.name), styles["SchoolTitle"]),
                Paragraph(
                    f"{category} | {school.classification.confidence.value.title()} confidence | "
                    f"{_pdf_text(school.institution.city)}, {_pdf_text(school.institution.state)}",
                    styles["BodyText"],
                ),
                Paragraph(
                    "ACT composite 25th-75th: "
                    f"{_pdf_text(_school_act_range(school))} "
                    "| "
                    "High-school GPA benchmark: "
                    f"{_pdf_text(school.high_school_gpa_benchmark or _GPA_UNAVAILABLE)}",
                    styles["Small"],
                ),
                Spacer(1, 6),
                _pdf_comparisons(school.comparisons, styles),
                Spacer(1, 8),
            ]
        )
        story.extend(_pdf_list("Strengths", school.strengths, styles))
        story.extend(_pdf_list("Gaps", school.gaps, styles))
        story.extend(_pdf_list("Missing data", school.unknowns, styles))
        story.extend(_pdf_list("Warnings", school.warnings, styles))
        story.extend(_pdf_list("Suggested actions", school.suggested_actions, styles))
        story.extend(_pdf_sources(school.source_references, styles))
        if index < len(report.schools) - 1:
            story.append(PageBreak())
    story.extend(_pdf_addendum(report, styles))
    document.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    return output.getvalue()


def export_college_report(
    report: CollegeReport,
    store: SessionFileStore,
    session_id: str,
    formats: set[ExportFormat],
    *,
    filename_stem: str | None = None,
) -> ExportResult:
    """Render requested formats and publish them through the storage boundary."""
    stem = filename_stem or f"college-report-{uuid4().hex[:8]}"
    rendered: list[tuple[ExportFormat, str, str, bytes]] = []
    if ExportFormat.PDF in formats:
        rendered.append((ExportFormat.PDF, f"{stem}.pdf", "application/pdf", render_pdf(report)))
    if ExportFormat.XLSX in formats:
        rendered.append(
            (
                ExportFormat.XLSX,
                f"{stem}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                render_excel(report),
            )
        )
    files = []
    for file_format, filename, media_type, content in rendered:
        path = store.write_file(session_id, filename, content)
        files.append(
            ExportedFile(
                format=file_format,
                filename=filename,
                path=str(path),
                media_type=media_type,
                size_bytes=len(content),
            )
        )
    return ExportResult(files=files)


def _college_list_sheet(workbook: Workbook, report: CollegeReport) -> None:
    sheet = workbook.create_sheet("College List")
    headers = [
        "Institution",
        "Classification",
        "Confidence",
        "User Entered",
        "City",
        "State",
        "Acceptance Rate",
        "ACT Composite 25th-75th",
        "High-School GPA Benchmark",
        "Cost of Attendance",
        "Average Net Price",
        "Graduation Rate",
        "Website",
        "Missing Data",
        "Warnings",
    ]
    rows = [
        [
            _excel_text(item.institution.name),
            item.classification.category.value.replace("_", " ").title(),
            item.classification.confidence.value.title(),
            "Yes" if item.user_entered else "No",
            _excel_text(item.institution.city),
            item.institution.state,
            item.institution.acceptance_rate,
            _act_range(item.institution.act_composite_25, item.institution.act_composite_75),
            item.high_school_gpa_benchmark or "Not available from current official dataset",
            item.institution.cost_of_attendance,
            item.institution.average_net_price,
            item.institution.graduation_rate,
            _excel_text(item.institution.website),
            _excel_text("; ".join(item.unknowns)),
            _excel_text("; ".join(item.warnings)),
        ]
        for item in report.schools
    ]
    _write_table_sheet(sheet, "CollegeListTable", headers, rows)
    for row_number, item in enumerate(report.schools, start=2):
        if item.institution.website:
            cell = sheet.cell(row_number, 13)
            cell.hyperlink = item.institution.website
            cell.style = "Hyperlink"
    for column in (7, 12):
        for cells in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
            for cell in cells:
                cell.number_format = "0.0%"
    for column in (10, 11):
        for cells in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
            for cell in cells:
                cell.number_format = '"$"#,##0'
    _classification_colors(sheet, f"B2:B{max(2, sheet.max_row)}")
    _set_widths(
        sheet,
        [30, 18, 12, 12, 16, 8, 15, 22, 36, 18, 17, 15, 28, 38, 38],
    )


def _student_supplied_colleges_sheet(workbook: Workbook, report: CollegeReport) -> None:
    sheet = workbook.create_sheet("Student-Supplied Colleges")
    headers = [
        "Student-Supplied Name",
        "Matched Institution",
        "Intended Major",
        "Classification",
        "Recommendation Rank Within Classification",
        "Category Applied Fit Threshold",
        "National Program Strength Rank",
        "Program Strength Rank Population",
        "National Program Strength Top Percent",
        "Program Strength Score",
        "Program Strength Confidence",
        "National Student-Major Fit Rank",
        "National Rank Population",
        "National Student-Major Fit Top Percent",
        "Fit Score",
        "Fit Confidence",
        "Match Status",
    ]
    school_by_name = {
        school.institution.name.casefold(): school
        for school in report.schools
        if school.user_entered
    }
    ranking_by_key = {
        (item.institution_name.casefold(), item.intended_major): item
        for item in report.student_supplied_rankings
    }
    rows: list[list[Any]] = []
    for supplied_name in report.student_profile.preferences.existing_schools:
        school = school_by_name.get(supplied_name.casefold())
        for major in report.student_profile.preferences.intended_majors:
            ranking = (
                ranking_by_key.get((school.institution.name.casefold(), major))
                if school is not None
                else None
            )
            rows.append(
                [
                    _excel_text(supplied_name),
                    _excel_text(school.institution.name) if school else None,
                    _excel_text(major),
                    (
                        school.classification.category.value.replace("_", " ").title()
                        if school
                        else None
                    ),
                    ranking.rank if ranking else None,
                    _threshold_for(report, major, school.classification.category)
                    if school
                    else None,
                    ranking.national_program_strength_rank if ranking else None,
                    ranking.national_program_strength_rank_total if ranking else None,
                    ranking.national_program_strength_top_percent if ranking else None,
                    float(ranking.program_strength_score) if ranking else None,
                    ranking.program_strength_confidence.value.title() if ranking else None,
                    ranking.national_rank if ranking else None,
                    ranking.national_rank_total if ranking else None,
                    ranking.national_fit_top_percent if ranking else None,
                    float(ranking.overall_score) if ranking else None,
                    ranking.confidence.value.title() if ranking else None,
                    "Matched" if school else "Not found in current official dataset",
                ]
            )
    _write_table_sheet(sheet, "StudentSuppliedCollegesTable", headers, rows)
    _classification_colors(sheet, f"D2:D{max(2, sheet.max_row)}", column="D")
    _set_widths(
        sheet,
        [30, 30, 24, 18, 24, 24, 28, 24, 22, 22, 22, 26, 22, 22, 12, 14, 34],
    )


def _major_rankings_sheet(workbook: Workbook, report: CollegeReport) -> None:
    sheet = workbook.create_sheet("Major Rankings")
    headers = [
        "Intended Major",
        "Institution",
        "Classification",
        "Recommendation Rank Within Classification",
        "Applied Fit Threshold",
        "National Program Strength Rank",
        "Program Strength Rank Population",
        "National Program Strength Top Percent",
        "Program Strength Score",
        "Program Strength Confidence",
        "National Student-Major Fit Rank",
        "National Rank Population",
        "National Student-Major Fit Top Percent",
        "Fit Score",
        "Fit Confidence",
        "Program Offered",
        "CIP Match Level",
        "Availability CIP6",
        "Ranking CIP4",
        "Mapped CIP Codes",
        "Academic Fit",
        "Major Fit",
        "Preferences",
        "Outcomes",
        "Holistic Alignment",
        "Missing Data",
        "Explanation",
    ]
    rows: list[list[Any]] = []
    for item in report.major_rankings:
        components = {component.name: component.score for component in item.components}
        rows.append(
            [
                _excel_text(item.intended_major),
                _excel_text(item.institution_name),
                item.category.value.replace("_", " ").title(),
                item.rank,
                item.applied_fit_threshold,
                item.national_program_strength_rank,
                item.national_program_strength_rank_total,
                item.national_program_strength_top_percent,
                float(item.program_strength_score),
                item.program_strength_confidence.value.title(),
                item.national_rank,
                item.national_rank_total,
                item.national_fit_top_percent,
                float(item.overall_score),
                item.confidence.value.title(),
                "Yes"
                if item.program_offered is True
                else "No"
                if item.program_offered is False
                else "Unknown",
                item.match_granularity,
                item.availability_cip_code,
                item.ranking_cip_code,
                ", ".join(item.cip_codes),
                components.get("Academic fit"),
                components.get("Major fit"),
                components.get("Student preferences"),
                components.get("Outcomes"),
                components.get("Holistic alignment"),
                _excel_text("; ".join(item.missing_inputs)),
                _excel_text(item.explanation),
            ]
        )
    _write_table_sheet(sheet, "MajorRankingsTable", headers, rows)
    for column in (5, 8, 9, 13, 14, 17, 21, 22, 23, 24, 25):
        for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
            for value in cell:
                value.number_format = "0.0"
    _classification_colors(sheet, f"C2:C{max(2, sheet.max_row)}", column="C")
    _set_widths(
        sheet,
        [
            24,
            30,
            18,
            24,
            20,
            28,
            24,
            22,
            22,
            22,
            26,
            22,
            22,
            12,
            14,
            16,
            14,
            16,
            14,
            24,
            13,
            13,
            13,
            13,
            16,
            38,
            55,
        ],
    )


def _addendum_sheet(workbook: Workbook, report: CollegeReport) -> None:
    sheet = workbook.create_sheet("Additional Qualified Colleges")
    headers = [
        "Intended Major",
        "Institution",
        "Classification",
        "Recommendation Rank Within Classification",
        "Applied Fit Threshold",
        "National Program Strength Rank",
        "Program Strength Rank Population",
        "National Program Strength Top Percent",
        "Program Strength Score",
        "Program Strength Confidence",
        "National Student-Major Fit Rank",
        "National Fit Rank Population",
        "National Student-Major Fit Top Percent",
        "Fit Score",
        "Fit Confidence",
    ]
    rows = [
        [
            _excel_text(item.intended_major),
            _excel_text(item.institution_name),
            item.category.value.replace("_", " ").title(),
            item.rank,
            item.applied_fit_threshold,
            item.national_program_strength_rank,
            item.national_program_strength_rank_total,
            item.national_program_strength_top_percent,
            float(item.program_strength_score),
            item.program_strength_confidence.value.title(),
            item.national_rank,
            item.national_rank_total,
            item.national_fit_top_percent,
            float(item.overall_score),
            item.confidence.value.title(),
        ]
        for item in report.addendum_rankings
    ]
    _write_table_sheet(sheet, "AdditionalQualifiedCollegesTable", headers, rows)
    _classification_colors(sheet, f"C2:C{max(2, sheet.max_row)}", column="C")
    _set_widths(
        sheet,
        [24, 32, 18, 30, 20, 28, 24, 22, 22, 22, 26, 22, 22, 12, 14],
    )


def _adaptive_thresholds_sheet(workbook: Workbook, report: CollegeReport) -> None:
    sheet = workbook.create_sheet("Adaptive Thresholds")
    headers = [
        "Intended Major",
        "Classification",
        "Mode",
        "Initial Threshold",
        "Applied Threshold",
        "Configured Floor",
        "Minimum Requested",
        "Exact Program Candidates",
        "Qualified Candidates",
        "Selected Candidates",
        "Addendum Candidates",
        "Threshold Relaxed",
    ]
    rows = [
        [
            _excel_text(item.intended_major),
            item.category.value.replace("_", " ").title(),
            item.threshold_mode.value.title(),
            float(item.initial_threshold),
            float(item.applied_threshold),
            float(item.adaptive_floor),
            item.minimum_requested,
            item.exact_program_candidates,
            item.qualified_candidates,
            item.selected_candidates,
            item.addendum_candidates,
            "Yes" if item.threshold_relaxed else "No",
        ]
        for item in report.category_thresholds
    ]
    _write_table_sheet(sheet, "AdaptiveThresholdsTable", headers, rows)
    _classification_colors(sheet, f"B2:B{max(2, sheet.max_row)}", column="B")
    _set_widths(sheet, [24, 18, 14, 18, 18, 18, 20, 24, 22, 20, 20, 18])


def _threshold_for(
    report: CollegeReport,
    major: str,
    category: Any,
) -> float | None:
    result = next(
        (
            item
            for item in report.category_thresholds
            if item.intended_major == major and item.category == category
        ),
        None,
    )
    return float(result.applied_threshold) if result else None


def _gap_analysis_sheet(workbook: Workbook, report: CollegeReport) -> None:
    sheet = workbook.create_sheet("Gap Analysis")
    headers = [
        "Institution",
        "Classification",
        "Measure",
        "Student Value",
        "School Benchmark",
        "Numeric Gap",
        "Status",
        "Note",
        "Source URL",
    ]
    rows = []
    for item in report.schools:
        for comparison in item.comparisons:
            rows.append(
                [
                    _excel_text(item.institution.name),
                    item.classification.category.value.replace("_", " ").title(),
                    comparison.measure,
                    _excel_text(comparison.student_value),
                    _excel_text(comparison.school_benchmark),
                    comparison.gap,
                    comparison.status.value.replace("_", " ").title(),
                    _excel_text(comparison.note),
                    _excel_text(comparison.sources[0].url if comparison.sources else None),
                ]
            )
    _write_table_sheet(sheet, "GapAnalysisTable", headers, rows)
    for row_number in range(2, sheet.max_row + 1):
        url = sheet.cell(row_number, 9).value
        if isinstance(url, str) and url:
            sheet.cell(row_number, 9).hyperlink = url.lstrip("'")
            sheet.cell(row_number, 9).style = "Hyperlink"
    _set_widths(sheet, [28, 18, 17, 16, 24, 13, 16, 45, 34])


def _student_profile_sheet(workbook: Workbook, report: CollegeReport) -> None:
    sheet = workbook.create_sheet("Student Profile")
    rows: list[list[Any]] = [
        ["Section", "Field", "Value", "Source"],
        [
            "Profile",
            "Applicant stage",
            report.student_profile.high_school.stage,
            "Confirmed profile",
        ],
        [
            "Profile",
            "Graduation year",
            report.student_profile.high_school.graduation_year,
            "Confirmed profile",
        ],
        [
            "Preferences",
            "Residence state",
            report.student_profile.preferences.residence_state,
            "Confirmed profile",
        ],
        [
            "Preferences",
            "Intended majors",
            _excel_text(", ".join(report.student_profile.preferences.intended_majors)),
            "Confirmed profile",
        ],
        [
            "Preferences",
            "Annual budget",
            report.student_profile.preferences.annual_budget,
            "Confirmed profile",
        ],
        [
            "Holistic context",
            "Résumé evidence review status",
            report.student_profile.holistic.review_status.value.replace("_", " ").title(),
            "Confirmed profile",
        ],
        [
            "Recommendation settings",
            "Threshold mode",
            report.student_profile.preferences.recommendation_settings.threshold_mode.value.title(),
            "Confirmed profile",
        ],
        [
            "Recommendation settings",
            "Initial fit threshold",
            report.student_profile.preferences.recommendation_settings.initial_fit_threshold,
            "Confirmed profile",
        ],
        [
            "Recommendation settings",
            "Adaptive floor",
            report.student_profile.preferences.recommendation_settings.adaptive_floor,
            "Confirmed profile",
        ],
        [
            "Recommendation settings",
            "Minimum results per category",
            report.student_profile.preferences.recommendation_settings.minimum_results_per_category,
            "Confirmed profile",
        ],
    ]
    for gpa in report.student_profile.academic.gpas:
        rows.append(
            [
                "Academics",
                f"{gpa.type.value.title()} {gpa.scope.value} GPA",
                f"{gpa.value}/{gpa.scale}",
                gpa.source.value,
            ]
        )
    for score in report.student_profile.academic.tests:
        label = score.test.value.upper()
        if score.section:
            label = f"{label} {score.section}"
        rows.append(["Testing", label, score.score, "Confirmed profile"])
    for course in report.student_profile.academic.courses:
        rows.append(
            [
                "Coursework",
                _excel_text(course.name or course.subject),
                _excel_text(course.grade.original if course.grade else course.status.value),
                course.source.value,
            ]
        )
    for theme in report.holistic_context.themes:
        rows.append(["Holistic context", "Theme", _excel_text(theme), "Confirmed résumé/activity"])
    _write_table_sheet(sheet, "StudentProfileTable", rows[0], rows[1:])
    for cell in sheet["C"]:
        if cell.row > 1 and sheet.cell(cell.row, 2).value == "Annual budget":
            cell.number_format = '"$"#,##0'
    _set_widths(sheet, [20, 30, 35, 22])


def _application_tracker_sheet(workbook: Workbook, report: CollegeReport) -> None:
    sheet = workbook.create_sheet("Application Tracker")
    headers = [
        "Institution",
        "Application Plan",
        "Deadline",
        "Status",
        "Fee",
        "Supplements",
        "Notes",
    ]
    rows = [
        [_excel_text(item.institution.name), "", None, "Not Started", None, "", ""]
        for item in report.schools
    ]
    _write_table_sheet(sheet, "ApplicationTrackerTable", headers, rows)
    if sheet.max_row >= 2:
        status_validation = DataValidation(
            type="list",
            formula1='"Not Started,In Progress,Submitted,Decision Received"',
            allow_blank=True,
        )
        sheet.add_data_validation(status_validation)
        status_validation.add(f"D2:D{sheet.max_row}")
        plan_validation = DataValidation(
            type="list",
            formula1='"Regular Decision,Early Action,Early Decision,Rolling"',
            allow_blank=True,
        )
        sheet.add_data_validation(plan_validation)
        plan_validation.add(f"B2:B{sheet.max_row}")
        for cell in sheet["C"][1:]:
            cell.number_format = "yyyy-mm-dd"
        for cell in sheet["E"][1:]:
            cell.number_format = '"$"#,##0'
    _set_widths(sheet, [30, 20, 14, 20, 12, 24, 40])


def _sources_sheet(workbook: Workbook, report: CollegeReport) -> None:
    sheet = workbook.create_sheet("Sources & Methodology")
    headers = ["Scope", "Source", "URL", "Source Date", "Methodology Version", "Notes"]
    rows: list[list[Any]] = [
        [
            "Dataset",
            report.dataset.source_name,
            _excel_text(report.dataset.source_url),
            report.dataset.release_date,
            report.methodology_version,
            f"Dataset version {report.dataset.version_id}",
        ]
    ]
    if report.fit_methodology_version:
        rows.append(
            [
                "Fit ranking",
                "Internal transparent fit methodology",
                "",
                None,
                report.fit_methodology_version,
                "Exact-program pool with per-category adaptive fit thresholds; national program "
                "strength precedes student fit. Not a commercial prestige ranking.",
            ]
        )
    for vintage in report.program_data_vintages:
        rows.append(["Program data", _excel_text(vintage), "", None, "", "Source vintage"])
    for warning in report.data_quality_warnings:
        rows.append(["Data quality", "Source-vintage warning", "", None, "", _excel_text(warning)])
    seen: set[tuple[str, str | None, Any]] = set()
    for item in report.schools:
        for source in item.source_references:
            key = (source.name, source.url, source.source_date)
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                [
                    _excel_text(item.institution.name),
                    _excel_text(source.name),
                    _excel_text(source.url),
                    source.source_date,
                    item.classification.methodology_version,
                    _excel_text(item.classification.explanation),
                ]
            )
    rows.append(
        [
            "Report",
            "Disclaimer",
            "",
            None,
            report.methodology_version,
            _excel_text(report.disclaimer),
        ]
    )
    _write_table_sheet(sheet, "SourcesMethodologyTable", headers, rows)
    for row_number in range(2, sheet.max_row + 1):
        url = sheet.cell(row_number, 3).value
        if isinstance(url, str) and url:
            sheet.cell(row_number, 3).hyperlink = url.lstrip("'")
            sheet.cell(row_number, 3).style = "Hyperlink"
        sheet.cell(row_number, 4).number_format = "yyyy-mm-dd"
    _set_widths(sheet, [24, 32, 42, 15, 20, 55])


def _write_table_sheet(sheet: Any, name: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    if not rows:
        sheet.append([None] * len(headers))
    table = Table(displayName=name, ref=f"A1:{_column_letter(len(headers))}{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=_THIN_GRAY)


def _classification_colors(sheet: Any, cell_range: str, *, column: str = "B") -> None:
    for text, color in (
        ("Safety Likely", _GREEN),
        ("Target", _BLUE),
        ("Reach", _RED),
        ("Insufficient Data", _GRAY),
    ):
        sheet.conditional_formatting.add(
            cell_range,
            FormulaRule(  # type: ignore[no-untyped-call]
                formula=[f'EXACT({column}2,"{text}")'],
                fill=PatternFill("solid", fgColor=color),
            ),
        )


def _set_widths(sheet: Any, widths: list[float]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[_column_letter(index)].width = width


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _excel_text(value: str | None) -> str | None:
    if value is None:
        return None
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def _act_range(low: int | None, high: int | None) -> str | None:
    if low is None or high is None:
        return None
    return f"{low}-{high}"


def _school_act_range(school: SchoolReport) -> str:
    return (
        _act_range(
            school.institution.act_composite_25,
            school.institution.act_composite_75,
        )
        or "Not available"
    )


def _pdf_summary(report: CollegeReport, styles: Any) -> PDFTable:
    counts = {category: 0 for category in ("safety_likely", "target", "reach", "insufficient_data")}
    for item in report.schools:
        counts[item.classification.category.value] += 1
    data: list[list[Any]] = [
        ["Schools", "Safety / Likely", "Target", "Reach", "Insufficient Data"],
        [
            len(report.schools),
            counts["safety_likely"],
            counts["target"],
            counts["reach"],
            counts["insufficient_data"],
        ],
    ]
    table = PDFTable(data, colWidths=[1.15 * inch] * 5)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_NAVY}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor(f"#{_LIGHT}")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B4C7E7")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table


def _pdf_thresholds(report: CollegeReport, styles: Any) -> list[Any]:
    if not report.category_thresholds:
        return []
    data: list[list[Any]] = [
        ["Major", "Category", "Initial", "Applied", "Floor", "Qualified", "Relaxed"]
    ]
    for item in report.category_thresholds:
        data.append(
            [
                Paragraph(_pdf_text(item.intended_major), styles["Small"]),
                item.category.value.replace("_", " ").title(),
                str(item.initial_threshold),
                str(item.applied_threshold),
                str(item.adaptive_floor),
                item.qualified_candidates,
                "Yes" if item.threshold_relaxed else "No",
            ]
        )
    table = PDFTable(
        data,
        colWidths=[
            1.45 * inch,
            1.15 * inch,
            0.65 * inch,
            0.7 * inch,
            0.6 * inch,
            0.75 * inch,
            0.65 * inch,
        ],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_NAVY}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (2, 1), (-1, -1), "CENTER"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor(f"#{_LIGHT}")],
                ),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E1F2")),
            ]
        )
    )
    return [
        Spacer(1, 12),
        Paragraph("Adaptive fit thresholds", styles["SchoolTitle"]),
        Paragraph(
            "Each category starts at the configured initial threshold and lowers independently "
            "until it reaches the requested option count or configured floor.",
            styles["Small"],
        ),
        Spacer(1, 6),
        table,
    ]


def _pdf_student_supplied_colleges(report: CollegeReport, styles: Any) -> list[Any]:
    supplied = report.student_profile.preferences.existing_schools
    if not supplied:
        return []
    school_by_name = {
        school.institution.name.casefold(): school
        for school in report.schools
        if school.user_entered
    }
    ranking_by_key = {
        (item.institution_name.casefold(), item.intended_major): item
        for item in report.student_supplied_rankings
    }
    data: list[list[Any]] = [
        [
            Paragraph(label, styles["TableHeader"])
            for label in (
                "Student-supplied college",
                "Major",
                "Category",
                "Category rank",
                "Threshold",
                "Program rank",
                "Fit rank",
            )
        ]
    ]
    for supplied_name in supplied:
        school = school_by_name.get(supplied_name.casefold())
        for major in report.student_profile.preferences.intended_majors:
            ranking = (
                ranking_by_key.get((school.institution.name.casefold(), major))
                if school is not None
                else None
            )
            data.append(
                [
                    Paragraph(_pdf_text(supplied_name), styles["Small"]),
                    Paragraph(_pdf_text(major), styles["Small"]),
                    (
                        school.classification.category.value.replace("_", " ").title()
                        if school
                        else "Not found"
                    ),
                    ranking.rank if ranking else "-",
                    _threshold_for(report, major, school.classification.category)
                    if school
                    else "-",
                    _pdf_rank(
                        ranking.national_program_strength_rank if ranking else None,
                        ranking.national_program_strength_rank_total if ranking else 0,
                        ranking.national_program_strength_top_percent if ranking else None,
                        styles,
                    ),
                    _pdf_rank(
                        ranking.national_rank if ranking else None,
                        ranking.national_rank_total if ranking else 0,
                        ranking.national_fit_top_percent if ranking else None,
                        styles,
                    ),
                ]
            )
    table = PDFTable(
        data,
        colWidths=[
            1.65 * inch,
            1.1 * inch,
            0.95 * inch,
            0.6 * inch,
            0.65 * inch,
            0.9 * inch,
            0.9 * inch,
        ],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_NAVY}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor(f"#{_LIGHT}")],
                ),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E1F2")),
            ]
        )
    )
    return [
        Spacer(1, 12),
        Paragraph("Student-supplied colleges", styles["SchoolTitle"]),
        Paragraph(
            "These colleges are retained and ranked separately before generated recommendations.",
            styles["Small"],
        ),
        Spacer(1, 6),
        table,
    ]


def _pdf_major_ranking_sections(report: CollegeReport, styles: Any) -> list[Any]:
    if not report.major_rankings:
        return []
    content: list[Any] = [Paragraph("Fit rankings by intended major", styles["SchoolTitle"])]
    content.append(
        Paragraph(
            "Each category applies its reported fit threshold, then prioritizes national program "
            "strength before student fit. Program strength uses free IPEDS/Scorecard program "
            "and outcome evidence; it is not a commercial prestige ranking. Remaining qualified "
            "colleges appear in the addendum.",
            styles["Small"],
        )
    )
    majors = list(dict.fromkeys(item.intended_major for item in report.major_rankings))
    for major_index, major in enumerate(majors):
        if major_index:
            content.append(PageBreak())
        content.append(Spacer(1, 8))
        content.append(Paragraph(_pdf_text(major), styles["SchoolTitle"]))
        data: list[list[Any]] = [
            [
                Paragraph(label, styles["TableHeader"])
                for label in (
                    "Strength rank",
                    "Fit rank",
                    "Category rank",
                    "Fit threshold",
                    "Institution",
                    "Category",
                    "Fit",
                    "Offered",
                    "Confidence",
                )
            ]
        ]
        for item in (value for value in report.major_rankings if value.intended_major == major):
            data.append(
                [
                    _pdf_rank(
                        item.national_program_strength_rank,
                        item.national_program_strength_rank_total,
                        item.national_program_strength_top_percent,
                        styles,
                    ),
                    _pdf_rank(
                        item.national_rank,
                        item.national_rank_total,
                        item.national_fit_top_percent,
                        styles,
                    ),
                    item.rank,
                    (
                        str(item.applied_fit_threshold)
                        if item.applied_fit_threshold is not None
                        else "-"
                    ),
                    Paragraph(_pdf_text(item.institution_name), styles["Small"]),
                    item.category.value.replace("_", " ").title(),
                    str(item.overall_score),
                    "Yes"
                    if item.program_offered is True
                    else "No"
                    if item.program_offered is False
                    else "Unknown",
                    item.confidence.value.title(),
                ]
            )
        table = PDFTable(
            data,
            colWidths=[
                0.8 * inch,
                0.8 * inch,
                0.65 * inch,
                0.65 * inch,
                1.45 * inch,
                0.9 * inch,
                0.5 * inch,
                0.65 * inch,
                0.75 * inch,
            ],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_NAVY}")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor(f"#{_LIGHT}")],
                    ),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E1F2")),
                ]
            )
        )
        content.append(table)
    return content


def _pdf_addendum(report: CollegeReport, styles: Any) -> list[Any]:
    if not report.addendum_rankings:
        return []
    data: list[list[Any]] = [
        [
            "Strength rank",
            "Institution",
            "Major",
            "Category",
            "Category rank",
            "Threshold",
            "Fit",
        ]
    ]
    for item in report.addendum_rankings:
        data.append(
            [
                _pdf_rank(
                    item.national_program_strength_rank,
                    item.national_program_strength_rank_total,
                    item.national_program_strength_top_percent,
                    styles,
                ),
                Paragraph(_pdf_text(item.institution_name), styles["Small"]),
                Paragraph(_pdf_text(item.intended_major), styles["Small"]),
                item.category.value.replace("_", " ").title(),
                item.rank,
                (
                    str(item.applied_fit_threshold)
                    if item.applied_fit_threshold is not None
                    else "-"
                ),
                str(item.overall_score),
            ]
        )
    table = PDFTable(
        data,
        colWidths=[
            0.85 * inch,
            1.8 * inch,
            1.2 * inch,
            1.0 * inch,
            0.7 * inch,
            0.65 * inch,
            0.5 * inch,
        ],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_NAVY}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor(f"#{_LIGHT}")],
                ),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E1F2")),
            ]
        )
    )
    return [
        PageBreak(),
        Paragraph("Addendum: additional qualified colleges", styles["SchoolTitle"]),
        Paragraph(
            "These colleges met their category's applied fit threshold but fell below the "
            "strongest programs selected for their admissions category.",
            styles["Small"],
        ),
        Spacer(1, 8),
        table,
    ]


def _pdf_comparisons(comparisons: list[Any], styles: Any) -> PDFTable:
    data: list[list[Any]] = [["Measure", "Student", "School benchmark", "Gap", "Status"]]
    for row in comparisons:
        data.append(
            [
                Paragraph(_pdf_text(row.measure), styles["Small"]),
                Paragraph(_pdf_text(row.student_value or "Unknown"), styles["Small"]),
                Paragraph(_pdf_text(row.school_benchmark or "Not published"), styles["Small"]),
                "" if row.gap is None else str(row.gap),
                row.status.value.replace("_", " ").title(),
            ]
        )
    table = PDFTable(
        data,
        colWidths=[1.1 * inch, 1.05 * inch, 1.9 * inch, 0.65 * inch, 1.15 * inch],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_NAVY}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{_LIGHT}")]),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E1F2")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _pdf_list(title: str, values: list[str], styles: Any) -> list[Any]:
    if not values:
        values = ["None identified from available data."]
    content: list[Any] = [Paragraph(f"<b>{escape(title)}</b>", styles["BodyText"])]
    content.extend(Paragraph(f"- {_pdf_text(value)}", styles["Small"]) for value in values)
    content.append(Spacer(1, 5))
    return content


def _pdf_sources(sources: list[Any], styles: Any) -> list[Any]:
    values = [
        f"{source.name}: {source.url or 'URL unavailable'}"
        + (f" ({source.source_date.isoformat()})" if source.source_date else "")
        for source in sources
    ]
    return _pdf_list("Sources", values, styles)


def _pdf_rank(
    rank: int | None,
    total: int,
    top_percent: Decimal | None,
    styles: Any,
) -> Any:
    if rank is None:
        return "Not ranked"
    percent = f"<br/>(top {top_percent}%)" if top_percent is not None else ""
    return Paragraph(f"{rank} of {total}{percent}", styles["Small"])


def _pdf_footer(canvas: Any, document: Any) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#666666"))
    canvas.drawString(0.55 * inch, 0.3 * inch, "US College Selection")
    canvas.drawRightString(7.95 * inch, 0.3 * inch, f"Page {document.page}")
    canvas.restoreState()


def _pdf_text(value: str) -> str:
    return escape(value.replace("\u2013", "-").replace("\u2014", "-"))
