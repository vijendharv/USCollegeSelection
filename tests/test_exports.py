from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from app.exporting import (
    _ordered_detail_schools,
    export_college_report,
    render_excel,
    render_pdf,
)
from app.models import ExportFormat
from app.ranking import rank_major_fits
from app.storage import LocalSessionFileStore
from tests.test_reporting import sample_report


def test_excel_has_required_sheets_tables_and_frozen_headers() -> None:
    workbook = load_workbook(BytesIO(render_excel(sample_report())))

    assert workbook.sheetnames == [
        "Student-Supplied Colleges",
        "College List",
        "Major Rankings",
        "Adaptive Thresholds",
        "Gap Analysis",
        "Student Profile",
        "Application Tracker",
        "Sources & Methodology",
    ]
    for sheet in workbook.worksheets:
        assert sheet.freeze_panes == "A2"
        assert len(sheet.tables) == 1
        assert not sheet.sheet_view.showGridLines


def test_excel_values_match_canonical_report() -> None:
    report = sample_report()
    workbook = load_workbook(BytesIO(render_excel(report)), data_only=False)
    sheet = workbook["College List"]

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert [row[0] for row in rows] == [item.institution.name for item in report.schools]
    assert [row[1] for row in rows] == [
        item.classification.category.value.replace("_", " ").title() for item in report.schools
    ]
    headers = [cell.value for cell in sheet[1]]
    assert "ACT Composite 25th-75th" in headers
    assert "High-School GPA Benchmark" in headers
    assert all(
        not (isinstance(cell.value, str) and cell.value.startswith("="))
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )


def test_student_supplied_colleges_are_separated_at_top() -> None:
    report = sample_report()
    workbook = load_workbook(BytesIO(render_excel(report)), data_only=False)
    sheet = workbook["Student-Supplied Colleges"]

    assert workbook.sheetnames[0] == "Student-Supplied Colleges"
    assert sheet["A2"].value == report.student_profile.preferences.existing_schools[0]
    assert sheet["Q2"].value == "Matched"


def test_excel_escapes_formula_injection_text() -> None:
    report = sample_report()
    report.schools[0].institution.name = '=HYPERLINK("https://bad.example")'

    workbook = load_workbook(BytesIO(render_excel(report)), data_only=False)
    cell = workbook["College List"]["A2"]

    assert cell.value == '\'=HYPERLINK("https://bad.example")'
    assert cell.data_type == "s"


def test_application_tracker_has_editable_status_validation() -> None:
    workbook = load_workbook(BytesIO(render_excel(sample_report())))
    sheet = workbook["Application Tracker"]

    validations = list(sheet.data_validations.dataValidation)
    assert len(validations) == 2
    assert any("Not Started" in (validation.formula1 or "") for validation in validations)


def test_pdf_contains_canonical_school_categories_and_missing_data() -> None:
    report = sample_report()
    reader = PdfReader(BytesIO(render_pdf(report)))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert "US College Selection Report" in text
    assert f"Methodology {report.methodology_version}" in text
    assert "Missing data" in text
    for item in report.schools:
        assert item.institution.name in text
        assert item.classification.category.value.replace("_", " ").title() in text


def test_pdf_has_page_numbers_and_ascii_dashes() -> None:
    reader = PdfReader(BytesIO(render_pdf(sample_report())))
    texts = [page.extract_text() or "" for page in reader.pages]

    assert len(texts) >= 5
    assert all(f"Page {index}" in text for index, text in enumerate(texts, start=1))
    assert all("\u2013" not in text and "\u2014" not in text for text in texts)


def test_pdf_has_clickable_navigation_destinations() -> None:
    report = sample_report()
    reader = PdfReader(BytesIO(render_pdf(report)))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert "Report navigation" in text
    assert "Back to report navigation" in text
    menu_links = reader.pages[0].get("/Annots", [])
    assert len(menu_links) >= len(report.schools) + 1
    assert all(page.get("/Annots") for page in reader.pages[1:])


def test_unranked_student_supplied_school_leads_detail_order() -> None:
    report = sample_report()

    ordered = _ordered_detail_schools(report)

    assert ordered[0].institution.name == report.student_profile.preferences.existing_schools[0]


def test_pdf_detail_order_follows_first_recommendation_table_appearance() -> None:
    report = sample_report()
    rankings, _ = rank_major_fits(report.student_profile, report.schools, [])
    ranking_by_id = {item.unit_id: item for item in rankings}
    expected = [report.schools[2], report.schools[0], report.schools[3], report.schools[1]]
    reordered = report.model_copy(
        update={"major_rankings": [ranking_by_id[item.institution.unit_id] for item in expected]}
    )

    ordered = _ordered_detail_schools(reordered)

    assert [item.institution.unit_id for item in ordered] == [
        item.institution.unit_id for item in expected
    ]


def test_export_service_writes_both_formats_through_session_store(tmp_path: Path) -> None:
    store = LocalSessionFileStore(tmp_path)
    session_id, session_path = store.create_session()

    result = export_college_report(
        sample_report(),
        store,
        session_id,
        {ExportFormat.PDF, ExportFormat.XLSX},
        filename_stem="college-report",
    )

    assert {file.format for file in result.files} == {ExportFormat.PDF, ExportFormat.XLSX}
    assert {file.filename for file in result.files} == {
        "college-report.pdf",
        "college-report.xlsx",
    }
    assert all(file.size_bytes > 0 for file in result.files)
    assert {path.name for path in session_path.iterdir()} == {
        "college-report.pdf",
        "college-report.xlsx",
    }
