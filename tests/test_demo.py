from __future__ import annotations

import json
from pathlib import Path
from typing import cast

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from app.cli import main
from app.demo import run_offline_demo

PROFILE = Path("examples/demo-student-profile.json")
FIXTURE = Path("tests/fixtures/scorecard/institutions.csv")


def test_offline_demo_builds_fixture_database_and_all_outputs(tmp_path: Path) -> None:
    result = run_offline_demo(
        profile_path=PROFILE,
        fixture_csv_path=FIXTURE,
        database_path=tmp_path / "data" / "college.duckdb",
        output_root=tmp_path / "output",
        schools_per_category=10,
    )

    assert result["status"] == "ok"
    assert result["offline"] is True
    assert result["database_built"] is True
    assert result["database_institutions"] == 6
    assert result["matching_institutions"] == 5
    assert result["reported_schools"] == 5
    files = [Path(path) for path in cast(list[str], result["files"])]
    assert {path.suffix for path in files} == {".pdf", ".xlsx", ".json"}
    assert all(path.is_file() for path in files)

    pdf = PdfReader(next(path for path in files if path.suffix == ".pdf"))
    pdf_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    workbook = load_workbook(next(path for path in files if path.suffix == ".xlsx"))
    report = json.loads(next(path for path in files if path.suffix == ".json").read_text())
    assert "Harvard University" in pdf_text
    assert workbook.sheetnames[0] == "College List"
    assert len(report["schools"]) == 5
    assert report["fit_methodology_version"] == "1.1"
    assert report["major_rankings"]
    assert (
        report["major_rankings"][0]["institution_name"]
        != "Arizona State University Campus Immersion"
    )
    assert "Major Rankings" in workbook.sheetnames


def test_offline_demo_reuses_existing_database(tmp_path: Path) -> None:
    database = tmp_path / "college.duckdb"
    output = tmp_path / "output"

    first = run_offline_demo(
        profile_path=PROFILE,
        fixture_csv_path=FIXTURE,
        database_path=database,
        output_root=output,
    )
    second = run_offline_demo(
        profile_path=PROFILE,
        fixture_csv_path=FIXTURE,
        database_path=database,
        output_root=output,
    )

    assert first["database_built"] is True
    assert second["database_built"] is False


def test_demo_requires_real_database_without_explicit_fixture(tmp_path: Path) -> None:
    result = main(
        [
            "demo",
            "--profile",
            str(PROFILE),
            "--database",
            str(tmp_path / "missing.duckdb"),
            "--output-dir",
            str(tmp_path / "output"),
        ]
    )

    assert result == 1


def test_demo_cli_runs_without_network_client(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_if_network_is_created(*args: object, **kwargs: object) -> None:
        raise AssertionError("demo must not create a network client")

    monkeypatch.setattr("app.cli.HttpClient", fail_if_network_is_created)
    result = main(
        [
            "demo",
            "--profile",
            str(PROFILE),
            "--fixture",
            str(FIXTURE),
            "--database",
            str(tmp_path / "college.duckdb"),
            "--output-dir",
            str(tmp_path / "output"),
        ]
    )

    payload = json.loads(capsys.readouterr().out)
    assert result == 0
    assert payload["offline"] is True
    assert payload["files"]
